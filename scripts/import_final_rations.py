import pandas as pd
import mysql.connector
from mysql.connector import Error
from datetime import datetime

# ── Config ──────────────────────────────────────────────────────
EXCEL_PATH = "/Users/olegzubarev/Desktop/FoodNutrion/финальный рационы март.xlsx"
DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "inbalance",
    "password": "inbalance_pass",
    "database": "inbalance",
}

# Sheet name -> (ration name, slug, calories, description, targetAudience)
RATIONS_DEF = {
    "1200 ккал": ("BALANCE 1200", "balance-1200", 1200,
     "Низкокалорийный рацион для похудения и детокса. Легкие блюда, богатые белком.",
     "Для тех, кто хочет сбросить вес и подтянуть фигуру."),
    "1500 ккал": ("BALANCE 1500", "balance-1500", 1500,
     "Сбалансированный рацион для комфортного снижения веса. Оптимальное соотношение КБЖУ.",
     "Для активных людей, которые хотят похудеть без голода."),
    "1800 ккал": ("BALANCE 1800", "balance-1800", 1800,
     "Рацион для поддержания формы и здорового образа жизни. Полноценное питание с достаточной калорийностью.",
     "Для тех, кто ведет активный образ жизни и занимается спортом."),
    "2000 ккал спорт": ("BALANCE 2000", "balance-2000", 2000,
     "Рацион для поддержания веса и высокой активности. Максимум энергии для тренировок и работы.",
     "Для спортсменов и людей с высокой физической нагрузкой."),
    "2500 ккал спорт": ("BALANCE 2500", "balance-2500", 2500,
     "Высококалорийный рацион для набора мышечной массы. Богатый белком и сложными углеводами.",
     "Для хардгейнеров и бодибилдеров в период набора массы."),
}

DAY_ORDER = {"Пн": 0, "Вт": 1, "Ср": 2, "Чт": 3, "Пт": 4, "Сб": 5, "Вс": 6}
DAY_ALIASES = {"ВТ": "Вт", "СР": "Ср", "пн": "Пн", "ПН": "Пн"}

MEAL_ORDER = {"Завтрак": 0, "Перекус": 1, "Обед": 2, "Ужин": 3}
MEAL_ALIASES = {"Доп": "Перекус", "перекус": "Перекус", "ужин": "Ужин", "Перекус": "Перекус"}


def normalize_day(day):
    if pd.isna(day):
        return None
    d = str(day).strip()
    d = DAY_ALIASES.get(d, d)
    return d if d in DAY_ORDER else None


def normalize_meal(meal):
    if pd.isna(meal):
        return None
    m = str(meal).strip()
    m = MEAL_ALIASES.get(m, m)
    return m if m in MEAL_ORDER else None


def clean_weight(val):
    if pd.isna(val):
        return ""
    v = str(val).strip()
    if v.lower() in ["1 порц.", "1 порц", "1 порция"]:
        return "1 порц."
    if v.endswith("г") and v[:-1].replace(".", "", 1).isdigit():
        return v
    if v.replace(".", "", 1).isdigit():
        return v + "г"
    return v


def clean_name(name):
    if pd.isna(name):
        return ""
    return str(name).strip().replace("\n", " / ")


def clean_number(val):
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.day
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_prices_from_sheet(ws):
    """Find price rows at the end of the sheet."""
    rows = list(ws.iter_rows(values_only=True))
    # Look for rows containing '1 день', '3 дня', etc.
    price_labels_row = None
    price_values_row = None
    for i in range(len(rows) - 1, -1, -1):
        row = rows[i]
        if any(isinstance(c, str) and "дн" in c.lower() for c in row if c):
            price_labels_row = row
            # The value row should be just before or after
            for j in range(i + 1, len(rows)):
                r = rows[j]
                if any(isinstance(c, str) and "р" in c for c in r if c):
                    price_values_row = r
                    break
            if not price_values_row and i > 0:
                for j in range(i - 1, -1, -1):
                    r = rows[j]
                    if any(isinstance(c, str) and "р" in c for c in r if c):
                        price_values_row = r
                        break
            break

    if not price_values_row:
        return [0, 0, 0, 0, 0]

    prices = []
    for val in price_values_row:
        if val and isinstance(val, str) and "р" in val:
            try:
                prices.append(int(val.replace("р", "").replace(" ", "").strip()))
            except ValueError:
                prices.append(None)
        elif val and isinstance(val, (int, float)):
            prices.append(int(val))
        else:
            prices.append(None)

    # Ensure we have 5 prices (1, 3, 5, 7, 14 days)
    while len(prices) < 5:
        prices.append(None)
    return prices[:5]


def main():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        # Clear everything
        cursor.execute("DELETE FROM dailyMealDishes")
        cursor.execute("DELETE FROM dailyMeals")
        cursor.execute("DELETE FROM rationDays")
        cursor.execute("DELETE FROM dishes")
        cursor.execute("DELETE FROM rations")
        conn.commit()

        dish_cache = {}  # name_lower -> dish_id
        ration_id = 0

        for sheet_name, (name, slug, calories, description, target) in RATIONS_DEF.items():
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ Sheet '{sheet_name}' not found, skipping")
                continue

            ration_id += 1
            sort_order = ration_id

            ws = wb[sheet_name]
            df = pd.DataFrame(ws.iter_rows(values_only=True))
            # Set first row as header
            if len(df) > 0:
                df.columns = df.iloc[0]
                df = df.iloc[1:].reset_index(drop=True)

            # Parse prices
            prices = parse_prices_from_sheet(ws)
            price1 = prices[0] or 0
            price3 = prices[1] or 0
            price5 = prices[2] or 0
            price7 = prices[3] or 0
            price14 = prices[4] or 0

            # Filter valid data rows
            valid_rows = []
            for _, row in df.iterrows():
                day = normalize_day(row.get("День"))
                meal = normalize_meal(row.get("Прием пищи"))
                dish_name = clean_name(row.get("Блюдо"))
                kcal = clean_number(row.get("Ккал"))
                if day and meal and dish_name and kcal is not None:
                    valid_rows.append(row)

            # Calculate average daily macros
            day_macros = {}
            for row in valid_rows:
                day = normalize_day(row.get("День"))
                protein = clean_number(row.get("Белки")) or 0
                fat = clean_number(row.get("Жиры")) or 0
                carbs = clean_number(row.get("Углеводы")) or 0
                kcal = clean_number(row.get("Ккал"))
                # Sanity check: fix obvious Excel typos where carbs are 10x+ too high
                if carbs > 150 and kcal and kcal > 0:
                    calc_carbs = (kcal - protein * 4 - fat * 9) / 4
                    if calc_carbs > 0:
                        carbs = round(calc_carbs, 1)
                if day:
                    if day not in day_macros:
                        day_macros[day] = [0.0, 0.0, 0.0]
                    day_macros[day][0] += protein
                    day_macros[day][1] += fat
                    day_macros[day][2] += carbs

            avg_protein = round(sum(v[0] for v in day_macros.values()) / max(len(day_macros), 1), 1)
            avg_fat = round(sum(v[1] for v in day_macros.values()) / max(len(day_macros), 1), 1)
            avg_carbs = round(sum(v[2] for v in day_macros.values()) / max(len(day_macros), 1), 1)

            # Insert ration
            cursor.execute(
                """
                INSERT INTO rations (id, name, slug, calories, protein, fat, carbs,
                    price1Day, price5Days, price7Days, price14Days,
                    description, targetAudience, sortOrder, isActive)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (ration_id, name, slug, calories, str(avg_protein), str(avg_fat), str(avg_carbs),
                 price1, price5, price7, price14,
                 description, target, sort_order, True),
            )
            conn.commit()

            print(f"\nImporting {name} ({len(valid_rows)} rows, prices={prices})...")

            # Track day occurrences to assign week
            day_occurrence = {}

            for row in valid_rows:
                day = normalize_day(row.get("День"))
                meal = normalize_meal(row.get("Прием пищи"))
                dish_name = clean_name(row.get("Блюдо"))
                weight = clean_weight(row.get("Вес, г"))
                protein = clean_number(row.get("Белки")) or 0
                fat = clean_number(row.get("Жиры")) or 0
                carbs = clean_number(row.get("Углеводы")) or 0
                kcal = clean_number(row.get("Ккал"))

                if not day or not meal or not dish_name:
                    continue
                if kcal is None:
                    continue

                # Sanity check: fix obvious Excel typos where carbs are 10x+ too high
                if carbs > 150 and kcal > 0:
                    calc_carbs = (kcal - protein * 4 - fat * 9) / 4
                    if calc_carbs > 0:
                        carbs = round(calc_carbs, 1)

                # Count day occurrence to determine week
                day_occurrence[day] = day_occurrence.get(day, 0) + 1
                occurrence = day_occurrence[day]
                # First 7 occurrences = week 1 (dayIndex 0-6), next 7 = week 2 (dayIndex 7-13)
                week = 0 if occurrence <= 7 else 1
                day_index = DAY_ORDER[day] + (week * 7)

                # Ensure dish exists
                name_key = dish_name.lower()
                if name_key not in dish_cache:
                    cursor.execute(
                        """
                        INSERT INTO dishes (name, calories, protein, fat, carbs, weight, isActive)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """,
                        (dish_name, int(kcal), str(protein), str(fat), str(carbs), weight, True),
                    )
                    conn.commit()
                    dish_cache[name_key] = cursor.lastrowid

                dish_id = dish_cache[name_key]

                # Ensure rationDay exists
                cursor.execute(
                    "SELECT id FROM rationDays WHERE rationId = %s AND dayIndex = %s",
                    (ration_id, day_index),
                )
                day_row = cursor.fetchone()
                if day_row:
                    ration_day_id = day_row["id"]
                else:
                    day_name_label = day if week == 0 else f"{day} (2)"
                    cursor.execute(
                        "INSERT INTO rationDays (rationId, dayIndex, dayName) VALUES (%s, %s, %s)",
                        (ration_id, day_index, day_name_label),
                    )
                    conn.commit()
                    ration_day_id = cursor.lastrowid

                # Create dailyMeal
                meal_sort = MEAL_ORDER[meal]
                cursor.execute(
                    "INSERT INTO dailyMeals (rationDayId, mealType, sortOrder) VALUES (%s, %s, %s)",
                    (ration_day_id, meal, meal_sort),
                )
                conn.commit()
                daily_meal_id = cursor.lastrowid

                # Add dish to meal
                cursor.execute(
                    "SELECT MAX(sortOrder) as max_sort FROM dailyMealDishes WHERE dailyMealId = %s",
                    (daily_meal_id,),
                )
                max_sort = cursor.fetchone()["max_sort"]
                dish_sort = (max_sort or 0) + 1

                cursor.execute(
                    "INSERT INTO dailyMealDishes (dailyMealId, dishId, weight, sortOrder) VALUES (%s, %s, %s, %s)",
                    (daily_meal_id, dish_id, weight, dish_sort),
                )
                conn.commit()

        print("\n✅ Import complete!")
        print(f"Total unique dishes: {len(dish_cache)}")

    except Error as e:
        print(f"❌ Database error: {e}")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


if __name__ == "__main__":
    main()
